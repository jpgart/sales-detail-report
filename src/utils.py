from analysis import charge_summary_new
import os
import re
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- DataFrame Utilities ---

def save_to_csv(df, file_path):
    """
    Save the DataFrame to a CSV file with UTF-8 encoding.
    """
    from column_name_map import rename_columns_for_readability
    df = rename_columns_for_readability(df)
    df.to_csv(file_path, index=False, encoding='utf-8-sig')

def load_data(file_path):
    """
    Load data from a CSV file into a DataFrame.
    """
    from column_name_map import normalize_dataframe_columns
    df = pd.read_csv(file_path, encoding='utf-8-sig')
    df = normalize_dataframe_columns(df)
    return df

def generate_report(df, report_path):
    """
    Generate a simple statistical report of the DataFrame and save it.
    """
    summary = df.describe()
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(str(summary))

# --- Exporter & Variety Utilities ---

def clean_exporter_name(lotdescr, lotid, mappings, specific_lotid_mappings):
    """
    Clean and standardize exporter names according to defined mappings and lotid overrides.
    """
    if lotid in specific_lotid_mappings:
        return specific_lotid_mappings[lotid]
    if pd.isna(lotdescr) or not str(lotdescr).strip():
        return "Unknown Exporter"
    lotdescr_upper = str(lotdescr).upper()
    for correct_name, misspellings_list in mappings.items():
        for item in misspellings_list:
            if item == lotdescr_upper:
                return correct_name
    parts = lotdescr_upper.split(' - ')
    if len(parts) > 1:
        potential_exporter_tag = parts[-1].strip()
        if potential_exporter_tag in mappings:
            return potential_exporter_tag
        for correct_name_map, misspellings_list_map in mappings.items():
            if potential_exporter_tag in misspellings_list_map:
                return correct_name_map
    for correct_name, misspellings_list in mappings.items():
        for misspelling in misspellings_list:
            if re.search(r'\b' + re.escape(misspelling) + r'\b', lotdescr_upper):
                return correct_name
            elif misspelling in lotdescr_upper and len(misspelling) > 3:
                return correct_name
    return "Unknown Exporter"

def get_variety(vrtyinvc_val, descr_val, known_varieties_set, known_exporter_tags_set, normalization_map):
    """
    Extract and normalize the grape variety name from 'vrtyinvc' or 'descr'.
    """
    if pd.notna(vrtyinvc_val):
        cleaned_vrty = str(vrtyinvc_val).strip().upper()
        normalized_vrty = normalization_map.get(cleaned_vrty, cleaned_vrty)
        if normalized_vrty in known_varieties_set:
            return normalized_vrty
    if pd.notna(descr_val):
        descr_upper = str(descr_val).strip().upper()
        parts = descr_upper.split(" - ", 1)
        if len(parts) > 0:
            potential_variety_from_descr = parts[0].strip()
            normalized_descr_vrty = normalization_map.get(potential_variety_from_descr, potential_variety_from_descr)
            if normalized_descr_vrty in known_varieties_set:
                if len(parts) > 1 and parts[1].strip() in known_exporter_tags_set:
                    return normalized_descr_vrty
                elif normalized_descr_vrty not in known_exporter_tags_set:
                    return normalized_descr_vrty
    return "Unknown Variety"

def extract_packaging_info(gwrproductdescr_val, pricepercase_val, detail_patterns, style_keywords):
    """
    Extract packaging style and detail from 'gwrproductdescr' and 'pricepercase'.
    """
    style = "Unknown"
    detail = "Unknown"
    if pd.isna(gwrproductdescr_val) or pd.isna(pricepercase_val):
        return style, detail
    gwr_upper = str(gwrproductdescr_val).upper()
    for pkg_detail, pattern in detail_patterns.items():
        if re.search(pattern, gwr_upper):
            detail = pkg_detail
            if "CLAM" in detail.upper():
                style = "Clam"
            elif "BAG" in detail.upper() or "POUCH" in detail.upper():
                style = "Bag"
            return style, detail
    for s_key, keywords in style_keywords.items():
        for keyword in keywords:
            if keyword in gwr_upper:
                style = s_key
                detail = f"{s_key} (Generic)"
                return style, detail
    return style, detail

def extract_retailer_name(descr_val, trxtype_val, invcicqnt_val, known_exporter_names_and_tags):
    """
    Extract the retailer name from 'descr' for sales transactions, excluding known exporter tags.
    """
    if pd.isna(descr_val) or not str(descr_val).strip():
        return "N/A"
    is_sales_transaction = (trxtype_val == 1) and (pd.notna(invcicqnt_val) and invcicqnt_val != 0)
    if is_sales_transaction:
        retailer_candidate = str(descr_val).strip()
        retailer_candidate_upper = retailer_candidate.upper()
        if " - " in retailer_candidate_upper:
            parts = retailer_candidate_upper.split(" - ", 1)
            if len(parts) > 1 and parts[1].strip() in known_exporter_names_and_tags:
                return "N/A (Likely Product/Exporter Info)"
        if retailer_candidate_upper in known_exporter_names_and_tags:
            return "N/A (Likely Product/Exporter Info)"
        return retailer_candidate if retailer_candidate else "Unknown Retailer"
    return "N/A"

def define_season(refdate):
    """
    Define the season based on the reference date.
    """
    if pd.isna(refdate):
        return "Undefined Season"
    if not isinstance(refdate, pd.Timestamp):
        try:
            refdate = pd.to_datetime(refdate)
        except ValueError:
            return "Invalid Date Format"
    if (pd.Timestamp('2023-11-01') <= refdate <= pd.Timestamp('2024-11-30')):
        return "2023-2024"
    elif (pd.Timestamp('2024-12-01') <= refdate <= pd.Timestamp('2025-11-30')):
        return "2024-2025"
    else:
        return "Undefined Season"

def get_exporter_country(exporter_name, exporter_country_map):
    """
    Get the country for an exporter using the provided mapping.
    """
    return exporter_country_map.get(exporter_name, "Unknown Country")

# --- Excel Export Function ---

def export_to_excel(dataframes_dict, sheet_descriptions_dict, output_excel_path):
    """
    Export a dictionary of DataFrames to an Excel file, creating a table of contents and applying formatting.
    """
    try:
        from column_name_map import rename_columns_for_readability
        # --- Ensure all numbers are exported as numbers, not strings ---
        def clean_numeric(df, sheet_name=None):
            df = df.copy()
            # Special handling for 'Sales Quantity' in Lot_Financial_Sum_All and per-exporter sheets
            if sheet_name is not None and sheet_name.startswith('Lot_Financial_Sum') and 'Sales Quantity' in df.columns:
                # Remove commas and convert to int
                df['Sales Quantity'] = pd.to_numeric(df['Sales Quantity'].astype(str).str.replace(',', ''), errors='coerce')
            for col in df.columns:
                # Remove $ and , and try to convert to float
                if df[col].dtype == object:
                    df[col] = df[col].astype(str).str.replace('$', '', regex=False).str.replace(',', '', regex=False)
                    try:
                        df[col] = pd.to_numeric(df[col], errors='coerce')
                    except Exception:
                        pass
            # Aplicar legibilidad de columnas antes de exportar
            df = rename_columns_for_readability(df)
            return df

        with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
            desc_sheet_name = "Sheet_Descriptions"
            if desc_sheet_name in writer.book.sheetnames:
                idx = writer.book.sheetnames.index(desc_sheet_name)
                writer.book.remove(writer.book[desc_sheet_name])
                ws_desc = writer.book.create_sheet(title=desc_sheet_name, index=idx)
            else:
                ws_desc = writer.book.create_sheet(title=desc_sheet_name, index=0)

            ws_desc.cell(row=1, column=1, value="Sheet Name").font = Font(bold=True)
            ws_desc.cell(row=1, column=2, value="Description").font = Font(bold=True)
            ws_desc.column_dimensions[get_column_letter(1)].width = 40
            ws_desc.column_dimensions[get_column_letter(2)].width = 120
            ws_desc.freeze_panes = 'A2'

            current_desc_row = 2

            # --- Custom sheet order as per user request ---
            preferred_order = [
                "Sheet_Descriptions",
                "Processed_Data",
                "Initial_Stock_All",
                "Initial_Stock_Summary",
                "Sales_Detail_By_Lotid",
                "Sales_Summary_By_Exporter",
                # "Retailer_Perf_Season",  # Eliminado
                # "Retailer_Perf_AllTime",  # Eliminado
                "Odd_Retailers",
                "Inventory_By_Exporter_FIFO",
                "Virtual_Inventory_Transactions",
                "Virtual_Inventory_Current_Summary",
                "Inventory_Summary_By_Exporter",
                "Lot_Financial_Sum_All",
                "Lot_Financial_Sum_Quintay",
                "Lot_Financial_Sum_MDT",
                "Lot_Financial_Sum_Agrovita",
                "Lot_Financial_Sum_Agrolatina",
                "Lot_Financial_Sum_Unknown_Expor",
                "Lot_Financial_Sum_Del_Monte",
                "Charge_Summary_Country",
                "Sales_Stddev_Consistency",
                "Stddev_Consistency_Summary",
                "Stddev_Consistency_By_Exporter",
                "Stddev_Consistency_By_Country"
            ]
            # Fuzzy match: for each preferred name, find the closest match in dataframes_dict
            from difflib import get_close_matches
            matched = set()
            ordered_sheet_names = []
            for name in preferred_order:
                matches = get_close_matches(name, dataframes_dict.keys(), n=1, cutoff=0.7)
                if matches:
                    ordered_sheet_names.append(matches[0])
                    matched.add(matches[0])
            # Add any extra sheets not in preferred order, sorted alphabetically
            extra_sheet_names = sorted([k for k in dataframes_dict.keys() if k not in matched])
            final_sheet_order = ordered_sheet_names + extra_sheet_names

            # Table of contents (Sheet_Descriptions)
            for sheet_name_full_key in final_sheet_order:
                safe_sheet_name_for_tab = "".join(c if c.isalnum() else "_" for c in sheet_name_full_key)[:31]
                base_desc_key = sheet_name_full_key
                prefixes = ["RetailerSales_", "ChargesPerLot_", "LotFinancialSum_Exp_", "DeducConsist_"]
                for prefix in prefixes:
                    if sheet_name_full_key.startswith(prefix):
                        base_desc_key = prefix.rstrip('_')
                        break
                else:
                    if "_" in sheet_name_full_key:
                        parts = sheet_name_full_key.split('_', 1)
                        base_desc_key = parts[0]
                        if len(parts) > 1:
                            first_part_of_suffix = parts[1].split('_')[0]
                            if parts[0] + "_" + first_part_of_suffix in sheet_descriptions_dict:
                                base_desc_key = parts[0] + "_" + first_part_of_suffix
                            elif parts[0] + "_" + parts[1] in sheet_descriptions_dict:
                                base_desc_key = parts[0] + "_" + parts[1]
                            elif base_desc_key == "Agrolatina" and "Specific_Charges" in sheet_name_full_key:
                                base_desc_key = "Agrolatina_Specific_Charges"
                description_text = sheet_descriptions_dict.get(base_desc_key, f"Data for {safe_sheet_name_for_tab}.")
                ws_desc.cell(row=current_desc_row, column=1, value=safe_sheet_name_for_tab)
                desc_cell = ws_desc.cell(row=current_desc_row, column=2, value=description_text)
                desc_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                current_desc_row += 1

            # Export all DataFrames as sheets, cleaning numbers, in the same order
            for sheet_name in final_sheet_order:
                df = dataframes_dict[sheet_name]
                safe_sheet_name = "".join(c if c.isalnum() else "_" for c in sheet_name)[:31]
                df_clean = clean_numeric(df, sheet_name=sheet_name)
                df_clean.to_excel(writer, sheet_name=safe_sheet_name, index=False)

            if "Sheet" in writer.book.sheetnames and writer.book["Sheet"] != ws_desc:
                std = writer.book["Sheet"]
                if std.max_row == 1 and std.max_column == 1 and std.cell(row=1, column=1).value is None:
                    writer.book.remove(std)
                    print("Removed default empty 'Sheet' after creating descriptions.")

            # Formatting pass (headers, column widths, freeze panes) in the same order
            for sheet_name_full in final_sheet_order:
                df_to_export = dataframes_dict[sheet_name_full]
                if df_to_export is not None and not df_to_export.empty:
                    safe_sheet_name = "".join(c if c.isalnum() else "_" for c in sheet_name_full)[:31]
                    if safe_sheet_name == desc_sheet_name:
                        continue
                    ws = writer.book[safe_sheet_name]
                    header_start_row = 1
                    for c_idx, col_name in enumerate(df_to_export.columns, 1):
                        cell = ws.cell(row=header_start_row, column=c_idx, value=col_name)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = Border(bottom=Side(style='thin'))
                    for r_idx, row_data in enumerate(dataframe_to_rows(df_to_export, index=False, header=False), header_start_row + 1):
                        for c_idx, value in enumerate(row_data, 1):
                            ws.cell(row=r_idx, column=c_idx, value=value)
                    for col_idx, column_title in enumerate(df_to_export.columns, 1):
                        column_letter = get_column_letter(col_idx)
                        current_width = 20
                        max_len = len(str(column_title)) + 2
                        for i in range(min(df_to_export.shape[0], 50)):
                            try:
                                if df_to_export.iloc[i, col_idx-1] is not None:
                                    max_len = max(max_len, len(str(df_to_export.iloc[i, col_idx-1])))
                            except: pass
                        adjusted_width = min(max(max_len + 2, current_width), 45)
                        ws.column_dimensions[column_letter].width = adjusted_width
                    if df_to_export.shape[0] > 0:
                        from openpyxl.cell.cell import MergedCell
                        freeze_row = header_start_row + 1
                        freeze_col = 2 if df_to_export.shape[1] > 1 else 1
                        cell_to_freeze = ws.cell(row=freeze_row, column=freeze_col)
                        # If the cell is a MergedCell, find the next non-merged cell in the row
                        if isinstance(cell_to_freeze, MergedCell):
                            # Try to find the next non-merged cell to the right
                            max_col = ws.max_column
                            found = False
                            for c in range(freeze_col, max_col + 1):
                                candidate = ws.cell(row=freeze_row, column=c)
                                if not isinstance(candidate, MergedCell):
                                    ws.freeze_panes = candidate
                                    found = True
                                    break
                            if not found:
                                # Fallback: freeze only the first row
                                ws.freeze_panes = ws.cell(row=freeze_row, column=1)
                        else:
                            ws.freeze_panes = cell_to_freeze
                    print(f"Sheet '{safe_sheet_name}' written to Excel.")
                else:
                    print(f"DataFrame for sheet '{sheet_name_full}' is empty or None, skipping.")
        print(f"Excel report generated: {output_excel_path}")
    except Exception as e:
        print(f"Error exporting to Excel: {e}")

# --- Blueprint & QC Markdown Generation ---

def generate_data_analysis_blueprint_v8(output_directory, script_version="8.0"):
    """
    Genera un blueprint Markdown documentando el flujo de análisis y procesamiento.
    """
    blueprint_content = f"""# Data Analysis Blueprint: Famus Extract V{script_version} Series

**Version:** {script_version} (Auto-generated by Python script)
**Date:** {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}
**Overall Objective:** To conduct a comprehensive analysis of sales data and deductions for grape shipments.
    1.  **Primary Output:** Detailed seasonal Excel reports with granular analyses for internal review, C-level data exploration, and exporter-specific insights. Each Excel file includes a "Sheet Descriptions" tab for clarity and features enhanced formatting (e.g., headers at row 1, bolding, adjusted column widths, freeze panes).
    2.  **Secondary Output:** Corresponding seasonal CSV files for each table generated in the Excel reports, stored in season-specific subdirectories.
    3.  **Integrated Output:** This Markdown blueprint, auto-generated with each run of the script to ensure documentation stays synchronized with the code.
    4.  **Integrated Output:** A Quality Control summary printed to the console upon completion of data processing and report generation for each season.

## 1. Data Sources

* **Primary Data:** Main sales and deductions data file (e.g., `JP Famus Report Original 05.15.25 - FAMOUS LOT DETAIL REPORT SA GRAPES 24-25.csv`).
* **Supporting Information (Embedded within the Python script):**
    * `EXPORTER_MAPPINGS`: Dictionary for correcting exporter name misspellings and mapping common tags.
    * `SPECIFIC_LOTID_EXPORTER_MAP`: Dictionary for `lotid`-to-exporter overrides.
    * `EXPORTER_COUNTRY_MAP`: Dictionary mapping cleaned exporter names to their countries.
    * `GRAPE_VARIETIES_LIST_FROM_USER` & `VARIETY_NORMALIZATION_MAP`: For standardizing grape variety names.
    * `PACKAGING_DETAIL_PATTERNS` & `PACKAGING_STYLE_KEYWORDS`: For extracting packaging information.
    * `AGROLATINA_SPECIFIC_CHARGES`: List of charges for specific analysis for Agrolatina.
    * Season Definitions:
        * "2023-2024": November 1, 2023 - November 30, 2024 (inclusive).
        * "2024-2025": December 1, 2024 - November 30, 2025 (inclusive).

## 2. Data Cleaning and Preprocessing Plan

* **Initial Load & Basic Cleaning:** Load data, convert date columns, clean and convert numeric columns (`pricepercase` to `pricepercase_cleaned`; `invcicqnt`, `rcptqnt`, `saleamt`, `chgamt`, `chgqnt`), handling NaNs.
* **Exporter Standardization:** Create `Exporter_Clean` column.
* **Country Mapping:** Create `Exporter_Country` column.
* **Variety Extraction:** Create `Variety` column.
* **Packaging Extraction:** Create `Packaging_Style` and `Packaging_Detail` columns from `gwrproductdescr` (for rows with `pricepercase` data).
* **Retailer Name Extraction:** Create `Retailer_Name` from `descr` for sales transactions (`trxtype == 1`, `invcicqnt != 0`).

## 3. Feature Engineering

* **`Season` Column:** Derived from `refdate`.
* **Commission & Advance Flags:** `Is_Advance`, `Is_ProducePay_Commission`, `Is_Retailer_Commission`.
* **Price Per Case Metrics:** `Price_Per_Case_invc` (`saleamt / invcicqnt`), `Price_Per_Case_rcpt` (`saleamt / rcptqnt`).
* **Readable Column Names:** A utility function (`rename_columns_for_readability`) is applied to all DataFrames before output to convert underscore_case or camelCase to "Title Case With Spaces".

## 4. Analysis Plan (Results generated per Season)

* **Lot Financial Summary (`calculate_lot_summary_with_exporter`):** Per-lot and exporter: Total Sales, Deductions (Excl. Advances), Advances, FOB Price, Advance % of FOB.
* **Pivoted Deductions (`analyze_deductions_by_category_per_lot`):** `lotid` & `Exporter_Clean` vs. `chargedescr` (sum of `chgamt`, excl. advances).
* **Deduction Exporter View (`analyze_deductions_exporter_view`):** Detailed deductions per `lotid`, `Exporter_Clean`, `chargedescr`, with `Total_Cases_Sold_Lot` and `Avg_Deduction_Per_Case`.
* **Ocean Freight:**
    * `get_detailed_ocean_freight_data`: Granular table.
    * `analyze_ocean_freight_summary_by_country`: Chile vs. Peru average comparison.
* **Retailer Sales by Exporter (`analyze_retailer_sales_by_exporter`):** Per exporter: summary by `Retailer_Name` (Total Cases, Sales, Avg Price). Includes "Exporter Clean" column.
* **Charges by Lot per Exporter (`analyze_charges_by_lot_per_exporter`):** Per exporter: pivot of all charges (`trxtype==2`) by `lotid` vs. `chargedescr`, with row/column totals. Includes "Exporter Clean" column.
* **Deduction Consistency (`analyze_deduction_consistency_by_charge`):** Per `chargedescr`: stats of `Avg_Deduction_Per_Case` across all relevant exporters for the season. All exporters for the season are listed.
* **Retailer Performance Detailed (`analyze_retailer_performance_detailed`):**
    # * `Retailer_Perf_Season`: (Eliminado)
    # * `Retailer_Perf_AllTime`: (Eliminado)
* **Odd Retailers (`analyze_odd_retailers`):** Sales with `invcicqnt != 0` but `saleamt` is zero/NaN.
* **Charge Characteristics:**
    * `analyze_fixed_vs_variable_charges`: Identifies charges often occurring without `chgqnt`.
    * `analyze_charge_rate_consistency`: Analyzes `chgamt/chgqnt` rates.
* **Fumigation Charges (`analyze_fumigation_charges`):** Checks application to non-Chilean exporters.
* **Commissions Overview (`analyze_commissions_overview`):** Summary of commission types.
* **Agrolatina Specific Charges (`analyze_agrolatina_specific_charges`):** Pivot table for predefined Agrolatina charges.
* **Data Quality Reporting (CSV/Excel Sheets):**
    * `DQ_Unknown_Exp`: Lots with "Unknown Exporter".
    * `DQ_Deduc_NoSales`: Lots with deductions but no sales.

## 5. Outputs (per Season)

### 5.1. Excel Reports
* **Filename:** `sales_analysis_report_[SeasonName].xlsx`.
* **First Sheet ("Sheet_Descriptions"):** Table of Contents listing all subsequent sheets and their detailed explanations (from `SHEET_DESCRIPTIONS` dictionary).
* **Data Sheets:** Headers start at Row 1. Includes tables for all analyses listed in section 4.
* **Formatting:** Headers are bold, centered, wrapped. Column widths adjusted (TOC Col A wider, data cols default/capped). Top row & first column frozen.

### 5.2. CSV Reports
* Stored in season-specific subdirectories (e.g., `CSV_Report_2023-2024/`).
* One CSV file per table generated in the Excel report.
* Filenames mirror Excel sheet names (e.g., `Deduc_Summary.csv`).
* Contain raw numeric data with readable column headers.

## 6. Quality Control Module (`run_quality_control_checks`)
* Executed after data processing and report generation for each season.
* Reads selected generated CSV files (or uses in-memory DataFrames) to perform validation checks.
* **Example Checks:**
    * **Totals Consistency:** Compares aggregated totals (e.g., total sales from 'Lot_Financial_Sum_All') against the sum of per-exporter totals (e.g., from 'LotFinancialSum_Exp_{{ExporterName}}' sheets) for key metrics like 'Total Sales', 'Total Cases Sold Invc'.
    * **Null Value Checks:** Verifies that critical identifier columns (e.g., 'Lotid', 'Exporter Clean', 'Retailer Name' in relevant tables) and key financial columns in summary tables do not contain unexpected nulls.
    * **Duplicate Checks:** Ensures no unintended duplicates for primary key combinations (e.g., 'Lotid' + 'Exporter Clean' should be unique in 'Lot_Financial_Sum_All').
    * **Formula Validation (Spot Checks):** Re-calculates key metrics like 'FOB Price' or 'Avg Deduction Per Case' for a sample of rows from relevant tables to ensure formulas are applied correctly.
    * **Logical Checks:**
        * Verifies that 'Total Advances' in 'Lot_Financial_Sum_All' are not negative.
        * Checks if 'Fumigation_Analysis' correctly identifies only Chilean exporters for fumigation charges (or flags discrepancies).
    * **Cross-Table Reconciliation (Example):** Verifies that the sum of 'Total Cases Sold Lot' from 'Deduc_Summary' for a given lot matches the 'Total Cases Sold Invc' for that lot in sales-related tables where appropriate.
* **Output:** Prints a summary of QC checks (pass/fail/warning) to the console for each season.

## 7. Future (To be handled by separate scripts or later versions)
* C-Level HTML Executive Summary Report.
* Interactive HTML Dashboards.
* Personalized PDF/HTML reports for individual exporters.
"""
    timestamp = pd.Timestamp.now().strftime('%Y%m%d_%H-%M-%S')
    blueprint_filename = os.path.join(output_directory, f"Data_Analysis_Blueprint_Famus_Extract_V{script_version.replace('.', '_')}_{timestamp}.md")
    try:
        with open(blueprint_filename, "w", encoding="utf-8") as f:
            f.write(blueprint_content)
        print(f"\nData Analysis Blueprint generated: {blueprint_filename}")
    except Exception as e:
        print(f"Error generating Data Analysis Blueprint: {e}")

def generate_qc_report_markdown(season_name, passed_all_qc, qc_issues_list, qc_log_messages, output_directory):
    """
    Genera un reporte Markdown para los resultados de los Quality Control Checks de una temporada.
    """
    report_filename = os.path.join(output_directory, f"QC_Report_{season_name.replace(' ', '_')}.md")
    report_content = f"# Quality Control Report for Season: {season_name}\n\n"
    report_content += f"**Date Generated:** {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
    report_content += "## Individual Check Results\n"
    for msg in qc_log_messages:
        report_content += f"- {msg}\n"
    report_content += "\n## Quality Control Summary\n"
    if passed_all_qc:
        report_content += f"✅ ALL CHECKS PASSED for season {season_name}!\n"
    else:
        report_content += f"⚠️ WARNING: Some QC checks FAILED for season {season_name}. Review issues:\n"
        for issue in qc_issues_list:
            report_content += f"  - {issue}\n"
    try:
        with open(report_filename, "w", encoding="utf-8") as f:
            f.write(report_content)
        print(f"\nQuality Control Report generated: {report_filename}")
    except Exception as e:
        print(f"Error generating QC Markdown report: {e}")

